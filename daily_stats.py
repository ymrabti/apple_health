"""
Apply daily aggregation and export to Excel file with three sheets:
1. Daily totals (Steps, Distance, Active Calories)
2. Weekly totals (Steps, Distance, Active Calories)
3. Daily statistics summary (Sum, Max, Min, Median, Average for each metric)
4. Date range of the data
"""

import sys
import statistics
from datetime import datetime, timezone, time, timedelta
import xml.etree.ElementTree as ET
from collections import defaultdict
from http.server import HTTPServer, BaseHTTPRequestHandler
import urllib.parse as urlparse
import urllib.request as urlrequest
import urllib.error as urlerror
import threading
import webbrowser
import signal
import time as systime
import shutil
import json
from openpyxl import Workbook
import keyring
from keyring.errors import KeyringError

APP_NAME = "health_dashboard"
BACKEND = "http://localhost:7384"
FRONTEND = "http://localhost:3762"
# ---- OAuth ----
SERVER_URL = f"{FRONTEND}/oauth/callback"
HEALTH_CHECK_ENDPOINT = f"{BACKEND}/api/health"
CRED_KEY = "jwt_token"

# ---- CONFIG ----
FILE = "export.xml"
# ----------------
tree = ET.parse(FILE)
root = tree.getroot()

# Global synchronization primitives for OAuth
AUTH_SUCCESS = threading.Event()
SERVER_READY = threading.Event()
SERVER_STATE = {"instance": None, "port": None}
CANCEL_EVENT = threading.Event()


def get_stored_token():
    """Retrieve stored JWT token from keyring."""
    return keyring.get_password(APP_NAME, CRED_KEY)


def store_token(token_back: str):
    """Store JWT token securely in keyring."""
    keyring.set_password(APP_NAME, CRED_KEY, token_back)


def clear_stored_token():
    """Clear stored JWT token from keyring."""
    keyring.delete_password(APP_NAME, CRED_KEY)


class CallbackHandler(BaseHTTPRequestHandler):
    """HTTP handler to process OAuth callback and extract JWT token."""

    def do_GET(self):
        """Handle GET request to extract JWT token from query parameters."""
        # extract JWT from query string: ?token=...

        query = urlparse.urlparse(self.path).query
        params = urlparse.parse_qs(query)
        token_oauth = params.get("token", [None])[0]
        if token_oauth:
            store_token(token_oauth)
            # Signal success to the main process
            AUTH_SUCCESS.set()
            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"Authentication successful. You can close this window.")
        else:
            self.send_response(400)
            self.end_headers()
            self.wfile.write(b"Authentication failed.")


def start_local_server(preferred_port: int = 11011):
    """Start a local HTTP server to handle OAuth callback (until shutdown).
    Tries preferred_port first, falls back to any free port.
    Signals readiness via SERVER_READY and records LISTEN_PORT.
    """
    try:
        httpd = HTTPServer(("127.0.0.1", preferred_port), CallbackHandler)
    except OSError:
        # Fallback to any available port
        httpd = HTTPServer(("127.0.0.1", 0), CallbackHandler)
    SERVER_STATE["instance"] = httpd
    SERVER_STATE["port"] = httpd.server_address[1]
    SERVER_READY.set()
    print(f"Local callback server listening on http://127.0.0.1:{SERVER_STATE['port']}")
    # Serve until `shutdown()` is called from the main thread.
    httpd.serve_forever(poll_interval=0.5)


def ensure_authenticated():
    """Ensure the user is authenticated and return the JWT token."""
    token_back = get_stored_token()
    if token_back:
        return token_back

    # launch local server in background (prefer 11011)
    server_thread = threading.Thread(
        target=start_local_server, kwargs={"preferred_port": 11011}, daemon=True
    )
    server_thread.start()

    # wait for server readiness to know the actual port
    if not SERVER_READY.wait(timeout=5):
        # could not start server; cancel
        if SERVER_STATE["instance"] is not None:
            try:
                SERVER_STATE["instance"].shutdown()
            except OSError:
                pass
        server_thread.join(timeout=2)
        print("❌ Failed to start local callback server.")
        return None

    # open browser to authenticate with the actual chosen port
    redirect_url = f"http://127.0.0.1:{SERVER_STATE['port']}/callback"
    auth_url = f"{SERVER_URL}?provider={redirect_url}"
    # Log the URL so the user can copy/paste it if needed
    print(
        "\nTo authenticate, open this URL in your browser (link valid for 5 minutes):",
        flush=True,
    )
    print(auth_url, "\n", flush=True)
    opened = webbrowser.open(auth_url)
    if not opened:
        print(
            "Browser did not open automatically. Please copy/paste the URL above.",
            flush=True,
        )

    # wait up to 5 minutes for authentication to complete; allow Ctrl+C to cancel
    deadline = systime.monotonic() + 30
    success = False
    try:
        while True:
            if AUTH_SUCCESS.is_set():
                success = True
                break
            if CANCEL_EVENT.is_set():
                raise KeyboardInterrupt
            remaining = deadline - systime.monotonic()
            if remaining <= 0:
                success = False
                break
            # poll in small intervals to allow signal handling
            AUTH_SUCCESS.wait(timeout=min(0.25, remaining))
    finally:
        # stop the local server regardless of outcome
        if SERVER_STATE["instance"] is not None:
            try:
                SERVER_STATE["instance"].shutdown()
            except OSError:
                pass
        server_thread.join(timeout=2)

    if not success:
        # Timed out or interrupted: cancel process
        return None

    return get_stored_token()


def validate_token(jwt_token: str, endpoint: str = HEALTH_CHECK_ENDPOINT) -> bool:
    """Validate the JWT by calling the health API endpoint.
    Returns True if the endpoint responds with HTTP 200, False otherwise.
    """
    try:
        url = f"{endpoint}"
        req = urlrequest.Request(
            url,
            headers={
                "Accept": "application/json",
                "Authorization": f"Bearer {jwt_token}",
            },
        )
        with urlrequest.urlopen(req, timeout=5) as resp:
            return resp.status == 200
    except (urlerror.URLError, urlerror.HTTPError) as e:
        print(f"❌ Token validation failed: {e}")
        return False


def parse_apple_health(start_dt, end_dt):
    """Parse Apple Health XML export and return daily summaries."""
    # Dummy implementation – replace with your parser
    date = start_dt
    data = []
    while date <= end_dt:
        data.append(
            {
                "date": date.strftime("%Y-%m-%d"),
                "steps": 10000,
                "distance": 7.2,
                "calories": 2100,
            }
        )
        date += timedelta(days=1)
    return data


def format_number(value, width=10):
    """
    Format a number with:
    - Thousand separator (.)
    - Decimal comma (,)
    - Left-padded with spaces to a total width
    """
    formatted = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return formatted.rjust(width, " ")


def finall_and_delete(output_path=None, make_backup=False):
    """Remove all Record elements from the XML tree and save to file.

    Args:
        output_path: Destination XML path. Defaults to sibling '<FILE>_cleaned.xml'.
        make_backup: If True and writing back to the original FILE, write a '.bak' backup first.

    Returns:
        The path to the saved XML file.
    """
    if output_path is None:
        if "." in FILE:
            base, ext = FILE.rsplit(".", 1)
            output_path = f"{base}_cleaned.{ext}"
        else:
            output_path = FILE + "_cleaned"

    # Remove all Record nodes
    records = list(root.findall(".//Record"))
    for record in records:
        root.remove(record)
    records = list(root.findall(".//ActivitySummary"))
    for record in records:
        root.remove(record)

    # Optional backup if overwriting original
    if make_backup and output_path == FILE:
        try:
            shutil.copyfile(FILE, FILE + ".bak")
        except OSError:
            pass

    # Pretty-print and write
    try:
        try:
            ET.indent(tree, space="  ")  # Python 3.9+
        except AttributeError:
            pass
        tree.write(output_path, encoding="utf-8", xml_declaration=True)
        print(f"✅ Saved cleaned XML to '{output_path}'")
        return output_path
    except OSError as e:
        print(f"❌ Failed to write cleaned XML: {e}")
        return output_path


def _post_json(url: str, payload: dict, jwt_token: str) -> bool:
    try:
        data = json.dumps(payload).encode("utf-8")
        req = urlrequest.Request(
            url,
            data=data,
            headers={
                "Accept": "application/json",
                "Content-Type": "application/json",
                "Authorization": f"Bearer {jwt_token}",
            },
            method="POST",
        )
        with urlrequest.urlopen(req, timeout=10) as resp:
            return 200 <= resp.status < 300
    except Exception as e:
        print(f"❌ POST {url} failed: {e}")
        return False


def _derive_export_date_str() -> str | None:
    try:
        elem = root.find(".//ExportDate")
        if elem is None:
            return None
        val = None
        if hasattr(elem, "attrib") and isinstance(elem.attrib, dict):
            val = elem.attrib.get("value") or elem.attrib.get("date")
        if not val:
            val = (elem.text or "").strip()
        if not val:
            return None
        # Try YYYY-MM-DD directly or extract from ISO
        if "T" in val:
            val = val.split("T")[0]
        if " " in val:
            val = val.split(" ")[0]
        date_str = val[:10]
        datetime.strptime(date_str, "%Y-%m-%d")
        return date_str
    except Exception:
        return None


def export_excel(_start, _end, jwt_token=None):
    """Export daily and weekly aggregated data to an Excel file."""
    # ---- Aggregate daily data ----
    # ---- Aggregate daily data ----
    initial_state = {
        "steps": 0,
        "distance": 0,
        "calories": 0,
        "basal_calories": 0,
        "flights": 0,
        "exercise": 0,
    }

    daily_data = defaultdict(lambda: initial_state)

    # ---- Aggregate weekly data ----
    weekly_data = defaultdict(lambda: initial_state)

    # Gather meta
    me_elem = root.find(".//Me")
    me_attrs = me_elem.attrib if me_elem is not None else {}
    export_date_str = _derive_export_date_str()
    # for record in root.findall(".//ActivitySummary"):
    #     print(record.attrib)

    record_dtypes = [
        "HKQuantityTypeIdentifierStepCount",
        "HKQuantityTypeIdentifierDistanceWalkingRunning",
        "HKQuantityTypeIdentifierActiveEnergyBurned",
        "HKQuantityTypeIdentifierBasalEnergyBurned",
        "HKQuantityTypeIdentifierFlightsClimbed",
        "HKQuantityTypeIdentifierAppleExerciseTime",
        "HKQuantityTypeIdentifierHeadphoneAudioExposure",
        "HKQuantityTypeIdentifierWalkingDoubleSupportPercentage",
        "HKQuantityTypeIdentifierWalkingSpeed",
        "HKQuantityTypeIdentifierWalkingStepLength",
        "HKQuantityTypeIdentifierWalkingAsymmetryPercentage",
        "HKQuantityTypeIdentifierAppleWalkingSteadiness",
        "HKCategoryTypeIdentifierSleepAnalysis",
        "HKCategoryTypeIdentifierHeadphoneAudioExposureEvent",
        "HKQuantityTypeIdentifierHeight",
        "HKQuantityTypeIdentifierBodyMass",
        "HKDataTypeSleepDurationGoal",
    ]

    # Map HealthKit record types to our daily_data keys
    aggregate_map = {
        "HKQuantityTypeIdentifierStepCount": "steps",
        "HKQuantityTypeIdentifierDistanceWalkingRunning": "distance",
        "HKQuantityTypeIdentifierActiveEnergyBurned": "calories",
        "HKQuantityTypeIdentifierBasalEnergyBurned": "basal_calories",
        "HKQuantityTypeIdentifierFlightsClimbed": "flights",
        "HKQuantityTypeIdentifierAppleExerciseTime": "exercise",
    }
    # Dynamic keys and preferred order
    metric_keys = list(dict.fromkeys(aggregate_map.values()))
    preferred_order = [
        "steps",
        "distance",
        "calories",
        "basal_calories",
        "flights",
        "exercise",
    ]
    ordered_keys = [k for k in preferred_order if k in metric_keys] + [
        k for k in metric_keys if k not in preferred_order
    ]

    # Reinitialize containers dynamically (overrides earlier static init)
    daily_data = defaultdict(lambda: {k: 0 for k in metric_keys})
    weekly_data = defaultdict(lambda: {k: 0 for k in metric_keys})

    def metric_label(key: str) -> str:
        mapping = {
            "steps": "Steps",
            "distance": "Distance (km)",
            "calories": "Active Calories (kcal)",
            "basal_calories": "Basal Calories (kcal)",
            "flights": "Flights",
            "exercise": "Exercise (minutes)",
        }
        return mapping.get(key, key.replace("_", " ").title())

    def format_cell(key: str, value: float):
        if key == "distance":
            return format_number(round(value, 2))
        return int(round(value))

    for record in root.findall(".//Record"):
        dtype = record.attrib.get("type")

        try:
            startdate = record.attrib.get("startDate")
            dt = datetime.fromisoformat(startdate.replace(" +", "+"))
        except ValueError:
            continue

        if _start <= dt <= _end:
            value_str = record.attrib.get("value", "0")
            try:
                value = float(value_str)
            except ValueError:
                continue

            day_key = dt.date()
            # Dynamic aggregation based on allowed dtypes
            if dtype in record_dtypes and dtype in aggregate_map:
                key = aggregate_map[dtype]
                daily_data[day_key][key] += value

    for day, data in daily_data.items():
        year, week, _ = day.isocalendar()
        week_key = f"{year}-W{week:02d}"
        for k in metric_keys:
            weekly_data[week_key][k] += data[k]

    # ---- Post data to backend (optional) ----
    if jwt_token:
        # 1) user infos
        user_infos_payload = {"exportDate": export_date_str, "attributes": me_attrs}
        _post_json(f"{BACKEND}/api/apple-health/user-infos", user_infos_payload, jwt_token)

        # 2) daily summaries
        # Build per-day summary objects to match DailySummary schema
        summaries_payload = []
        for day, metrics in daily_data.items():
            def _int(v):
                try:
                    return int(round(float(v)))
                except Exception:
                    return 0

            def _dec(v):
                try:
                    return round(float(v), 4)
                except Exception:
                    return 0.0

            item = {
                "date": day.isoformat(),
                "steps": _int(metrics.get("steps", 0)),
                "flights": _int(metrics.get("flights", 0)),
                "distance": _dec(metrics.get("distance", 0)),
                "active": _dec(metrics.get("calories", 0)),
                "basal": _dec(metrics.get("basal_calories", 0)),
                "exercise": _dec(metrics.get("exercise", 0)),
            }
            if export_date_str:
                item["exportDate"] = export_date_str
            summaries_payload.append(item)

        _post_json(
            f"{BACKEND}/api/apple-health/daily-summaries",
            {"summaries": summaries_payload},
            jwt_token,
        )

        # 3) activity summaries (filtered by date range if possible)
        act_summaries = []
        for rec in root.findall(".//ActivitySummary"):
            attrs = dict(rec.attrib)
            dc = attrs.get("dateComponents")
            try:
                if dc:
                    d = datetime.strptime(dc, "%Y-%m-%d").date()
                    if not (_start.date() <= d <= _end.date()):
                        continue
            except Exception:
                pass
            act_summaries.append(attrs)
        _post_json(
            f"{BACKEND}/api/apple-health/activity-summaries",
            {"exportDate": export_date_str, "summaries": act_summaries},
            jwt_token,
        )

    # ---- Create Excel file ----
    wb = Workbook()

    # --- Daily Sheet ---
    daily_sheet = wb.active
    daily_sheet.title = "Daily Totals"
    daily_sheet.append(["Date"] + [metric_label(k) for k in ordered_keys])
    for day in sorted(daily_data.keys()):
        data = daily_data[day]
        row = [day.isoformat()] + [format_cell(k, data[k]) for k in ordered_keys]
        daily_sheet.append(row)

    # --- Weekly Sheet ---
    weekly_sheet = wb.create_sheet(title="Weekly Totals")
    weekly_sheet.append(["Week"] + [metric_label(k) for k in ordered_keys])
    for week in sorted(weekly_data.keys()):
        data = weekly_data[week]
        row = [week] + [format_cell(k, data[k]) for k in ordered_keys]
        weekly_sheet.append(row)

    # --- Daily Statistics Sheet ---
    stats_sheet = wb.create_sheet(title="Daily Stats Summary")
    stats_sheet.append(
        [
            "Metric",
            "Sum",
            "Max",
            "Day Hit Max",
            "Min",
            "Day Hit Min",
            "Median",
            "Average",
        ]
    )

    # Prepare lists and corresponding dates dynamically
    all_dates = list(daily_data.keys())
    series_by_key = {k: [data[k] for data in daily_data.values()] for k in ordered_keys}

    def get_day_of_value(lst, dates, value):
        """Return the first date corresponding to the value"""
        for v, d in zip(lst, dates):
            if v == value:
                return d.isoformat()
        return ""

    metrics = [(metric_label(k), series_by_key[k], all_dates) for k in ordered_keys]

    for name, lst, dates in metrics:
        max_val = max(lst)
        min_val = min(lst)
        stats_sheet.append(
            [
                name,
                format_number(round(sum(lst), 2)),
                format_number(round(max_val, 2)),
                get_day_of_value(lst, dates, max_val),
                format_number(round(min_val, 2)),
                get_day_of_value(lst, dates, min_val),
                format_number(round(statistics.median(lst), 2)),
                format_number(round(statistics.mean(lst), 2)),
            ]
        )
    stats_sheet.append(
        [
            "Date Range",
            "Start Date",
            min(daily_data.keys()).isoformat(),
            "End Date",
            max(daily_data.keys()).isoformat(),
        ]
    )
    output_xlsx = f"activity_summaries/{_start.date()}-{_end.date()}.xlsx"
    # ---- Save Excel ----
    wb.save(output_xlsx)
    print(f"Daily, weekly, and daily stats summary exported to '{output_xlsx}'")


# --- 1️⃣ Parse CMD Arguments ---
if len(sys.argv) != 3:
    print("Usage: python daily_stats.py <start_date> <end_date>")
    print("Example: python daily_stats.py 2025-08-01 2025-08-31")
    sys.exit(1)


def _sigint_handler(_signum, _frame):
    CANCEL_EVENT.set()


try:
    # install Ctrl+C handler
    try:
        signal.signal(signal.SIGINT, _sigint_handler)
    except ValueError:
        # In some environments signals may not be configurable; ignore
        pass
    start_date = datetime.combine(
        datetime.strptime(sys.argv[1], "%Y-%m-%d").date(), time.min, tzinfo=timezone.utc
    )
    end_date = datetime.combine(
        datetime.strptime(sys.argv[2], "%Y-%m-%d").date(), time.max, tzinfo=timezone.utc
    )

    token = ensure_authenticated()
    if CANCEL_EVENT.is_set():
        print("\nOperation cancelled by user.")
        sys.exit(1)
    if not token:
        print("⚠️ Authentication timed out (5 minutes). Process cancelled.")
        sys.exit(1)
    # Validate token with local API before proceeding; if invalid, start OAuth
    if not validate_token(token):
        print("⚠️ Stored token invalid. Starting OAuth re-authentication…")
        try:
            print("Clearing stored token.")
            clear_stored_token()
        except KeyringError:
            pass
        token = ensure_authenticated()
        if CANCEL_EVENT.is_set():
            print("\nOperation cancelled by user.")
            sys.exit(1)
        if not token:
            print("⚠️ Authentication timed out (5 minutes). Process cancelled.")
            sys.exit(1)
        if not validate_token(token):
            print("❌ Authentication failed: token invalid after re-authentication.")
            sys.exit(1)
    summaries = parse_apple_health(start_date, end_date)
    export_excel(start_date, end_date, token)
except ValueError:
    print("❌ Dates must be in YYYY-MM-DD format")
    sys.exit(1)
except KeyboardInterrupt:
    # Ensure server is shut down if running and exit cleanly
    if SERVER_STATE["instance"] is not None:
        try:
            SERVER_STATE["instance"].shutdown()
        except OSError:
            pass
    print("\nOperation cancelled by user.")
    sys.exit(1)
